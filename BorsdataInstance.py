# -*- coding: utf-8 -*-
"""
Created on Thu Aug  8 11:49:59 2019

@author: Jack
"""


from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait

from time import sleep
import time,datetime
import pandas as pd
import os
import shutil

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

currentDirPath = os.getcwd()
foldername = os.path.basename(currentDirPath)

if('Dokument\\Aktier' not in currentDirPath):
    raise SystemError('The current directory path does not contain Dokument\\Aktier, your current directory path is: '+currentDirPath)
if('Aktier' != foldername):
    raise SystemError('The current foldername is not "Aktier", your current folder name is: ' + foldername)
    


class BorsdataInstance:
    URL = 'https://borsdata.se/'
    DOWNLOADPATH = os.getcwd() + '\\Data'
    def __init__(self,user_credentials_path: str):
        if type(user_credentials_path) is not str:
            raise TypeError('Expected a string as user_credentials_path but got a '+ str(type(user_credentials_path))+'.')
        if not os.path.exists(user_credentials_path):
            raise IOError(user_credentials_path + ' does not exist.')
        self.user_credentials_path = user_credentials_path
        self.loggedIn = False
        self.markets = Markets()
        self.markets.addSweden('all')
        self.excludedIndustries = []
    def Login(self):
        with open(self.user_credentials_path) as f:
            lineList = f.readlines()
        username = lineList[0]
        password = lineList[1]
        
        options = webdriver.ChromeOptions()
        prefs= {"browser.download.dir":BorsdataInstance.DOWNLOADPATH}
        options.add_experimental_option("prefs",prefs)
        
        self.driver = webdriver.Chrome(options = options)
        self.driver.implicitly_wait(10)

        self.driver.get(BorsdataInstance.URL)
        
        sleep(1)
        loginbutton = self.driver.find_element_by_id('LogInOpenButton')
        loginbutton.click()
        
        
        usernameForm = self.driver.find_element_by_id('login_UserName')
        passwordForm = self.driver.find_element_by_id('login_Password')
        
        
        sleep(1)
        usernameForm.click()
        usernameForm.send_keys(username)
        sleep(1)
        passwordForm.click()
        passwordForm.send_keys(password)
        
        loginbutton_final = self.driver.find_element_by_id('submitLogin')
        loginbutton_final.click()
        sleep(3)
        self.loggedIn = True
        return
    
    def updateAll(self):
        validStrategies = ['MagicFormula','AcqMu']
        dataList = {}
        for strategy in validStrategies:
            df = self.updateStrategy(strategy)
            dataList[strategy] = df
        return dataList
        
    def updateStrategy(self,strategy):
        validStrategies = ['MagicFormula','AcqMu']
        if strategy not in validStrategies:
            raise ValueError('Valid strategies are MagicFormula and AcqMu, you entered: ' + str(strategy))
        
        self.driver.get(BorsdataInstance.URL+'export')
        sleep(2)
        retrieveSavebutton = self.driver.find_element_by_class_name('savings-button')
        retrieveSavebutton.click()
        sleep(0.5)
        print('Using strategy: '+ strategy)
        StrategyFilter = self.driver.find_element_by_link_text(strategy)
        StrategyFilter.click()
        sleep(3)
        
        
        self.driver.find_element_by_css_selector('markets-selection-button').click()
        sleep(2)
        self.clickMarkets()

        sleep(2)
        self.driver.find_element_by_class_name('markets-selection-window__buttons__finalize').click()
        sleep(1)
        self.driver.find_element_by_css_selector('industries-selection-button').click()
        sleep(2)
        self.removeIndustries(strategy = strategy)
        sleep(2)
        self.driver.find_element_by_css_selector('button[ng-click="industriesCtrl.submit()"]').click()
        sleep(2)
        exportButton = self.driver.find_element_by_class_name('export-button')
        exportButton.click()
        
        
        def_downloadPath = os.path.expanduser('~')+ '\\Downloads'
        downloadName =def_downloadPath+'\\'+  'Borsdata_' + str(datetime.datetime.now().date())+'.xlsx'
        movepath = BorsdataInstance.DOWNLOADPATH +'\\'+ 'Borsdata_' +strategy+ '_Scraped.xlsx'
        
        max_time = 15
        time_waited = 0
        while (not os.path.exists(downloadName)) and time_waited < max_time:
            time.sleep(1)
            time_waited += 1
        
        shutil.move(downloadName, movepath)
        #self.driver.quit()
        sleep(1)
        
        df = BorsdataInstance.BorsdataProcessScraped(strategy = strategy)
        BorsdataInstance.BorsdataUpdateHistAndCurr(df, strategy = strategy)
    
        return df
        
    def BorsdataProcessScraped(*,MarketCap_minimum = 500,strategy):
        path = 'Data\Borsdata_'+strategy+'_Scraped.xlsx'
        if strategy == 'MagicFormula':
            df = pd.read_excel(path)
            df = df.drop(0)
            
            df = df[pd.notnull(df['Magic'])]
            df = df[pd.notnull(df['ROC'])]
            df = df[pd.notnull(df['EBIT/EV (%)'])]
            df = df[df['Börsvärde'] > MarketCap_minimum]
            df.index = range(len(df.index))
            
            df['ROC_rank'] = df['ROC'].rank(ascending = False)
            df['EBIT_rank'] = df['EBIT/EV (%)'].rank(ascending = False)
            df['Magic_rank'] = (df['ROC_rank']+df['EBIT_rank']).rank()
            
            df = df.sort_values(by = 'Magic_rank')
            
            df.insert(1,'Magic_rank',df.pop('Magic_rank'))
            df.insert(2,'ROC_rank',df.pop('ROC_rank'))
            df.insert(3,'EBIT_rank',df.pop('EBIT_rank'))
            df.insert(4,'Senaste rapport',df.pop('Info.6'))
            df.insert(5,'Rapport datum',df.pop('Info.7'))
        elif strategy == 'AcqMu':
            df = pd.read_excel(path)
            df = df.drop(0)
            
            df = df[pd.notnull(df['EV/OP'])]
            df = df[pd.notnull(df['EV'])]
            df = df[pd.notnull(df['OP kassaf.'])]
            df = df[df['OP kassaf.'] > 0]
            df = df[df['Börsvärde']> MarketCap_minimum]
            df.index = range(len(df.index))
            
            df['AcqMu rank'] = df['EV/OP'].rank(ascending = True)
            
            df = df.sort_values(by = 'AcqMu rank')
            
            df.insert(1,'AcqMu rank',df.pop('AcqMu rank'))
            
        else:
            raise ValueError('Valid strategies are MagicFormula and AcqMu')
        return df
    def BorsdataUpdateHistAndCurr(df,*,strategy):
        histPath = 'Data\Borsdata_'+ strategy +'_Historik.xlsx'
        currPath = 'Data\Borsdata_'+ strategy +'_Current.xlsx'
        curr_date = str(datetime.datetime.now().date())
        try:
            wb = load_workbook(histPath)
        except FileNotFoundError:
            wb = Workbook()
        try: #Removes curr_date sheet if it exists otherwise creates a new one.
            wb.remove(wb[curr_date])
        except Exception:
            print()
        finally:
            wb.create_sheet(curr_date)
            wb.active = wb[curr_date]
        ws = wb.active
        
        for r in dataframe_to_rows(df,index = False, header = True):
            ws.append(r)
        for cell in ws[1]: #Adds pandas style to first row (headers)
            cell.style = 'Pandas'
        for col in ws.columns: #Fixes column width to an appropriate size.
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try: #Avoids error on empty cells.
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length+2)*1.1
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(histPath)
        
        wb_current = load_workbook(histPath)    #Loads the historical data file,
        sheets = wb_current.sheetnames          #removes all sheets except curr_date
        for s in sheets:                        #and saves it as Borsdata_strategy_Current.xlsx
            if s != curr_date:
                sheet = wb[s]
                wb.remove(sheet)
        wb.save(currPath)
        return
    
      
    def moveTo(self,site: str):
        self.driver.get(BorsdataInstance.URL + site)
        return
    def clickMarkets(self):
        
        self.driver.find_element_by_css_selector('button[ng-click = "marketsCtrl.deselectAll()"]').click()
        #Deselects all markets
        checkboxes = self.driver.find_elements_by_class_name('items-group-box__label')
        for country in self.markets.Countries:
            for currCheckbox in checkboxes:
                if country in currCheckbox.get_property('innerText'):
                    currCheckbox.click()
        subcheckboxes = self.driver.find_elements_by_class_name('items-group-box__sub-items__li')
        for swemarket in self.markets.Sweden:
            for currCheckbox in subcheckboxes[0:8]:
                if swemarket in currCheckbox.get_property('innerText'):
                    currCheckbox.click()
                    print('Adding market: Sweden,' + currCheckbox.get_property('innerText'))
        if 'Sverige' in self.markets.Sweden:
            checkboxes[0].click()
            print('Adding market: Sweden')
        for normarket in self.markets.Norway:
            for currCheckbox in subcheckboxes[8:14]:
                if normarket in currCheckbox.get_property('innerText'):
                    currCheckbox.click()
                    print('Adding market: Norway,' + currCheckbox.get_property('innerText'))
        if 'Norge' in self.markets.Norway:
            checkboxes[1].click()
            print('Adding market: Norway')
        for finmarket in self.markets.Finland:
            for currCheckbox in subcheckboxes[14:20]:
                if finmarket in currCheckbox.get_property('innerText'):
                    currCheckbox.click()
                    print('Adding market: Finland,' + currCheckbox.get_property('innerText'))
        if 'Finland' in self.markets.Finland:
            checkboxes[2].click()
            print('Adding market: Finland')
        for denmarket in self.markets.Denmark:
            for currCheckbox in subcheckboxes[20:26]:
                if denmarket in currCheckbox.get_property('innerText'):
                    currCheckbox.click()
                    print('Adding market: Denmark,' + currCheckbox.get_property('innerText'))
        if 'Danmark' in self.markets.Denmark:
            checkboxes[3].click()
            print('Adding market: Denmark')
    def removeIndustries(self,strategy):
        self.driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div/div[2]/div[1]/button[1]").click()
        industriesList = self.driver.find_elements_by_class_name('items-group-box__label')
        industryToRemove = []
        if strategy == 'MagicFormula':
            industryToRemove = ['Finans & Fastighet']
        elif strategy == 'AcqMu':
            industryToRemove = ['Energi','Finans & Fastighet']
        
        for industry in industriesList:
            for toRemove in industryToRemove:
                if toRemove in industry.get_property('innerText'):
                    print('Removing industry:' + industry.get_property('innerText'))
                    industry.click()
                    
    def downloadBolagsdata(self,companyName,ticker):
        self.driver.get(BorsdataInstance.URL)
        #Searches for company
        wait = WebDriverWait(self.driver,10)
        self.driver.find_element_by_css_selector('#CompListTop > input').send_keys(companyName)
        #Clicks onto company site
        sleep(2)
        companylist = self.driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div/div/div[4]/div/div/div[1]/div/table')
        companylist.click()
        
        #Clicks onto export page
        sleep(2)
        exportPage = self.driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[2]/div[2]/ui-view/div[1]/div/div/button')
        exportPage.click()
        sleep(2)
        self.driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[2]/div[2]/ui-view/div[1]/div/div/button')
        #Click export button
        self.driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[2]/div[2]/ui-view/div/div[3]/div[2]/button').click()
        
        def_downloadPath = os.path.expanduser('~')+ '\\Downloads\\'
        filename = ticker + '-' + companyName + '.xlsx'
        downloadName = def_downloadPath + filename
        movepath  = BorsdataInstance.DOWNLOADPATH + '\\Bolagsdata\\'+filename
        
        max_time = 15
        time_waited = 0
        while (not os.path.exists(downloadName)) and time_waited < max_time:
            time.sleep(1)
            time_waited += 1
        
        shutil.move(downloadName,movepath)
        return(movepath)
    def extractAcqMudata(self,path):
        wb = pd.read_excel(path,sheet_name = None)
        
        Year_sheet = wb['Year']
        Report_series = Year_sheet['Report']
        wanted_reports = ('EV/OP','OP-marginal','Antal Aktier','EBITDA','EV/EBITDA','Direktavkastning','Utdelning')
        i = 0
        indices = []
        for report in wanted_reports:
            indices.append(Report_series[Report_series == report].index[0])
            i += 1
            
        fund_data = Year_sheet.iloc[indices]
        fund_data.drop(columns = ['Unnamed: 1'],inplace = True)
        fund_data.drop(fund_data.columns[[-1,]],axis = 1, inplace = True)
        print(fund_data)
        EV = fund_data.iloc[4][1:]*fund_data.iloc[3][1:]
        Market_cap = fund_data.iloc[6][1:] / fund_data.iloc[5][1:] * fund_data.iloc[2][1:] *100
        print(EV,Market_cap)
        Price_sheet = wb['PriceMonth']
        
        year = 2010
        startdate = str(year) + '-03-01'
        enddate = str(year) + '-03-31'
        dates = pd.DataFrame()
        while year <= datetime.datetime.now().year-1:
            dates = pd.DataFrame.append(dates,Price_sheet[(Price_sheet['Date']>=startdate)&(Price_sheet['Date'] <= enddate)])
            year += 1
            startdate = str(year) + '-03-01'
            enddate = str(year) + '-03-31'
        return(dates)
class Markets:
    
    def __init__(self):
        self.Countries = []
        self.Sweden = []
        self.Norway = []
        self.Finland = []
        self.Denmark = []
        
    def addCountry(self,newCountry:str):
        if newCountry not in self.Countries and newCountry != 'all':
            self.Countries.append(newCountry)
        if newCountry == 'all':
            self.Countries = ['USA','Kanada','Storbrittanien','Tyskland','Frankrike',
                              'Spanien','Portugal','Italien','Schweiz','Belgien','Nederländerna',
                              'Polen','Estland','Lettland','Litauen','Island']
        return
    def removeCountry(self,toRemove):
        if toRemove in self.Countries:
            self.Countries.remove(toRemove)
        if toRemove == 'all':
            self.Countries = []
    def addSweden(self,market:str):
        if market not in self.Sweden and market != 'all':
            self.Sweden.append(market)
        if market == 'all':
            self.Sweden = ['Sverige']
    def removeSweden(self,market:str):
        if market in self.Sweden and market != 'all':
            self.Sweden.remove(market)
        if market == 'all':
            self.Sweden = []
            
    def addNorway(self,market:str):
        if market not in self.Norway and market != 'all':
            self.Norway.append(market)
        if market == 'all':
            self.Norway = ['Norge']
    def removeNorway(self,market:str):
        if market in self.Norway and market != 'all':
            self.Norway.remove(market)
        if market == 'all':
            self.Norway = []
    
    def addFinland(self,market:str):
        if market not in self.Finland and market != 'all':
            self.Finland.append(market)
        if market == 'all':
            self.Finland = ['Finland']
    def removeFinland(self,market:str):
        if market in self.Finland and market != 'all':
            self.Finland.remove(market)
        if market == 'all':
            self.Finland = []

    def addDenmark(self,market:str):
        if market not in self.Denmark and market != 'all':
            self.Denmark.append(market)
        if market == 'all':
            self.Denmark = ['Danmark']
    def removeDenmark(self,market:str):
        if market in self.Norway and market != 'all':
            self.Denmark.remove(market)
        if market == 'all':
            self.Denmark = []

bd = BorsdataInstance('user_info.txt')
#bd.extractAcqMudata(path)
bd.Login()
#bd.downloadBolagsdata('AarhusKarlshamn','AAK')

#bd.getBolagsdata('AarhusKarlshamn')
data = bd.updateAll()